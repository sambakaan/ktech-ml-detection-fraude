"""
Gestion de l'historique des prédictions (session + fichier persistant) et génération
des rapports exportables (CSV, PDF, Excel).
"""
import io
import os

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def charger_historique_persistant(history_path: str) -> pd.DataFrame:
    """Charge l'historique des prédictions déjà enregistré sur disque (le cas échéant)."""
    if os.path.exists(history_path):
        try:
            return pd.read_csv(history_path)
        except Exception:
            return pd.DataFrame()
    return pd.DataFrame()


def ajouter_ligne_historique(ligne: dict, history_path: str):
    """Ajoute une prédiction à l'historique en session ET au fichier persistant sur disque.

    Limite : sur Streamlit Community Cloud, le disque est réinitialisé à chaque redéploiement /
    redémarrage du conteneur. Cet historique est donc fiable pendant la durée de vie de
    l'instance en cours, mais n'est pas une base de données à long terme.
    """
    st.session_state["historique_predictions"].append(ligne)

    os.makedirs(os.path.dirname(history_path), exist_ok=True)
    df_ligne = pd.DataFrame([ligne])
    fichier_existe = os.path.exists(history_path)
    df_ligne.to_csv(history_path, mode="a", header=not fichier_existe, index=False)


def vider_historique(history_path: str):
    st.session_state["historique_predictions"] = []
    if os.path.exists(history_path):
        os.remove(history_path)


@st.cache_data
def generer_pdf_historique(df: pd.DataFrame) -> bytes:
    """Génère un rapport PDF (paysage) résumant l'historique des prédictions.

    Mis en cache par contenu : si le tableau filtré n'a pas changé depuis le dernier
    rendu, on ne relance pas la génération ReportLab à chaque interaction sans rapport
    (ex : changer un filtre non lié sur la page ne doit pas recalculer le PDF)."""
    buffer = io.BytesIO()
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        title="Historique des prédictions - Détection de fraude bancaire",
    )

    story = [
        Paragraph("Historique des prédictions — Détection de fraude bancaire", styles["Title"]),
        Paragraph(
            f"Généré le {pd.Timestamp.now().strftime('%d/%m/%Y à %H:%M')} — {len(df)} transaction(s) analysée(s)",
            styles["Normal"],
        ),
        Spacer(1, 0.5 * cm),
    ]

    colonnes_affichees = [
        "Horodatage", "Type de transaction", "Status operation", "Localisation",
        "Montant", "Statut prédit", "Probabilité Fraude (%)",
    ]
    colonnes_affichees = [c for c in colonnes_affichees if c in df.columns]
    data = [colonnes_affichees] + df[colonnes_affichees].astype(str).values.tolist()

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#153B58")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B7C6D6")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF1FB")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def _styliser_feuille_excel(feuille):
    """Applique l'habillage KTech Solutions (en-tête navy, texte blanc gras) à une feuille,
    et ajuste la largeur des colonnes à leur contenu."""
    remplissage_entete = PatternFill(start_color="153B58", end_color="153B58", fill_type="solid")
    police_entete = Font(color="FFFFFF", bold=True)
    for cellule in feuille[1]:
        cellule.fill = remplissage_entete
        cellule.font = police_entete

    for colonne in feuille.columns:
        longueur = max((len(str(c.value)) for c in colonne if c.value is not None), default=10)
        feuille.column_dimensions[get_column_letter(colonne[0].column)].width = min(longueur + 2, 40)


def generer_excel_historique(df: pd.DataFrame, resume: dict | None = None) -> bytes:
    """Génère un classeur Excel multi-feuilles : détail des transactions + résumé."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Transactions", index=False)
        _styliser_feuille_excel(writer.sheets["Transactions"])
        if resume:
            pd.DataFrame(list(resume.items()), columns=["Indicateur", "Valeur"]).to_excel(
                writer, sheet_name="Résumé", index=False
            )
            _styliser_feuille_excel(writer.sheets["Résumé"])
    buffer.seek(0)
    return buffer.getvalue()
